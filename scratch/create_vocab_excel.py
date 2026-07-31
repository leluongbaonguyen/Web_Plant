import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Sheet 1: NHẬP DỮ LIỆU ────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "📥 Nhập Dữ Liệu"

HEADERS = [
    ("word",               "TỪ TIẾNG ANH *",          "apple",          "BẮT BUỘC. Chữ thường, không dấu."),
    ("meaning",            "NGHĨA TIẾNG VIỆT *",      "quả táo",        "BẮT BUỘC. Ngắn gọn, chuẩn tiểu học."),
    ("ipa",                "PHIÊN ÂM IPA",             "/ˈæp.əl/",       "Tra tại oxfordlearnersdictionaries.com"),
    ("vietnamesePhonetic", "ĐỌC TIẾNG VIỆT",          "Áp-pờ-lơ",       "Phiên âm hướng dẫn đọc cho bé."),
    ("type",               "LOẠI TỪ",                 "Danh từ",        "Danh từ | Động từ | Tính từ | Số đếm"),
    ("image",              "EMOJI / HÌNH ẢNH",        "🍎",             "Emoji Unicode hoặc để trống."),
    ("hint",               "MẸO NHỚ",                 "💡 Táo đỏ ngọt!", "Bắt đầu bằng 💡. Tối đa 80 ký tự."),
    ("example",            "CÂU VÍ DỤ TIẾNG ANH",    "I eat an apple.", "Ngắn, phù hợp lứa tuổi 4-10."),
    ("exampleVi",          "CÂU VÍ DỤ TIẾNG VIỆT",   "Tôi ăn quả táo.", "Dịch tương ứng 1-1 với câu EN."),
    ("level",              "CẤP ĐỘ",                  "L1",             "L1 | L2 | L3 | L4"),
    ("category",           "MÃ CHỦ ĐỀ",              "L1-U05",         "Định dạng: [Level]-U[Số] VD: L2-U03"),
    ("audioUrl",           "LINK ÂM THANH",           "",               "URL file MP3. Để trống = dùng TTS."),
]

HEADER_COLOR   = "1E293B"  # dark slate
SUBHEAD_COLOR  = "0F766E"  # teal
REQUIRED_COLOR = "DC2626"  # red
SAMPLE_COLOR   = "F0FDF4"  # light green
HINT_COLOR     = "EFF6FF"  # light blue

thin = Side(style="thin", color="CBD5E1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell_style(cell, bg, fg="FFFFFF", bold=False, size=10, wrap=False, align="left"):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(color=fg, bold=bold, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = border

# ── Title row ─────────────────────────────────────────────────────────────────
ws1.merge_cells("A1:L1")
title_cell = ws1["A1"]
title_cell.value = "📚 CHRONOFLOW PREMIUM – FILE MẪU NHẬP DỮ LIỆU TỪ VỰNG TIẾNG ANH CHO BÉ"
title_cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
title_cell.font = Font(color="FBBF24", bold=True, size=14, name="Calibri")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 36

# ── Sub-title ─────────────────────────────────────────────────────────────────
ws1.merge_cells("A2:L2")
sub = ws1["A2"]
sub.value = "Hướng dẫn: Điền dữ liệu từ hàng 6 trở xuống. Cột có dấu * là BẮT BUỘC. Sau khi điền xong, copy toàn bộ dữ liệu và dán vào ô Import Wizard trên hệ thống."
sub.fill = PatternFill("solid", fgColor="0F766E")
sub.font = Font(color="FFFFFF", size=10, name="Calibri")
sub.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[2].height = 28

# ── Column headers ────────────────────────────────────────────────────────────
COL_WIDTHS = [18, 22, 20, 22, 14, 12, 32, 36, 36, 8, 12, 28]

for col_idx, (field, label, sample, hint) in enumerate(HEADERS, start=1):
    col_letter = get_column_letter(col_idx)
    ws1.column_dimensions[col_letter].width = COL_WIDTHS[col_idx - 1]

    # Row 3: field name (technical)
    c3 = ws1.cell(row=3, column=col_idx, value=f'"{field}"')
    cell_style(c3, "334155", "94A3B8", bold=False, size=8, align="center")
    ws1.row_dimensions[3].height = 18

    # Row 4: display label
    is_req = "*" in label
    c4 = ws1.cell(row=4, column=col_idx, value=label)
    bg = REQUIRED_COLOR if is_req else SUBHEAD_COLOR
    cell_style(c4, bg, "FFFFFF", bold=True, size=10, align="center")
    ws1.row_dimensions[4].height = 24

    # Row 5: hint
    c5 = ws1.cell(row=5, column=col_idx, value=hint)
    cell_style(c5, "1E3A5F", "93C5FD", bold=False, size=8, wrap=True, align="left")
    ws1.row_dimensions[5].height = 32

# ── Sample data rows ──────────────────────────────────────────────────────────
SAMPLE_DATA = [
    ["apple",        "quả táo",           "/ˈæp.əl/",          "Áp-pờ-lơ",        "Danh từ", "🍎", "💡 Táo đỏ ngọt ngào mỗi ngày!",          "I eat a red apple.",            "Tôi ăn một quả táo đỏ.",        "L1", "L1-U05", ""],
    ["banana",       "quả chuối",         "/bəˈnæn.ə/",        "Bơ-na-na",         "Danh từ", "🍌", "💡 Chuối vàng cung cấp năng lượng!",      "Monkeys love bananas.",          "Khỉ rất thích ăn chuối.",       "L1", "L1-U05", ""],
    ["elephant",     "con voi",           "/ˈel.ə.fənt/",      "E-lơ-phơn-tơ",     "Danh từ", "🐘", "💡 Con voi có cái vòi rất dài!",          "The elephant is very big.",      "Con voi rất to lớn.",           "L2", "L2-U01", ""],
    ["rainbow",      "cầu vồng",          "/ˈreɪnboʊ/",        "Rên-bâu",          "Danh từ", "🌈", "💡 7 màu cầu vồng sau cơn mưa rào!",     "Look at the rainbow!",          "Nhìn cầu vồng kìa!",            "L3", "L3-U02", ""],
    ["astronaut",    "phi hành gia",      "/ˈæstrənɔːt/",      "Át-strơ-nót",      "Danh từ", "🧑‍🚀", "💡 Phi hành gia bay vào vũ trụ xa xôi!", "The astronaut walks in space.",  "Phi hành gia đi bộ trong không gian.", "L4", "L4-U01", ""],
    ["run",          "chạy",              "/rʌn/",             "Răn",              "Động từ", "🏃", "💡 Chạy thật nhanh như tên lửa!",         "I run every morning.",           "Tôi chạy mỗi buổi sáng.",       "L1", "L1-U07", ""],
    ["happy",        "vui vẻ",            "/ˈhæp.i/",          "Hép-pi",           "Tính từ", "😊", "💡 Khi được quà là vui vẻ nhất!",         "I am happy today.",              "Hôm nay tôi rất vui vẻ.",       "L1", "L1-U08", ""],
    ["mountain",     "núi",               "/ˈmaʊn.tɪn/",       "Maun-tin",         "Danh từ", "⛰️", "💡 Núi cao chọc trời mây trắng!",         "The mountain is very high.",     "Ngọn núi rất cao.",             "L3", "L3-U03", ""],
    ["computer",     "máy tính",          "/kəmˈpjuː.tər/",   "Cơm-piu-tơ",       "Danh từ", "💻", "💡 Máy tính giúp học và chơi game!",      "I use a computer every day.",   "Tôi dùng máy tính mỗi ngày.",   "L3", "L3-U01", ""],
    ["dream",        "giấc mơ",           "/driːm/",           "Đờ-rim",           "Danh từ", "💭", "💡 Hãy mơ thật lớn và cố gắng!",          "I have a big dream.",            "Tôi có một giấc mơ lớn.",       "L4", "L4-U10", ""],
]

for row_idx, row_data in enumerate(SAMPLE_DATA, start=6):
    alt_bg = "F8FAFC" if row_idx % 2 == 0 else "FFFFFF"
    for col_idx, value in enumerate(row_data, start=1):
        c = ws1.cell(row=row_idx, column=col_idx, value=value)
        c.fill = PatternFill("solid", fgColor=alt_bg)
        c.font = Font(size=10, name="Calibri", color="1E293B")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = border
    ws1.row_dimensions[row_idx].height = 22

# Add 30 empty data rows
for row_idx in range(16, 46):
    alt_bg = "F8FAFC" if row_idx % 2 == 0 else "FFFFFF"
    for col_idx in range(1, 13):
        c = ws1.cell(row=row_idx, column=col_idx, value="")
        c.fill = PatternFill("solid", fgColor=alt_bg)
        c.font = Font(size=10, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border
    ws1.row_dimensions[row_idx].height = 22

ws1.freeze_panes = "A6"

# ── Sheet 2: HƯỚNG DẪN ───────────────────────────────────────────────────────
ws2 = wb.create_sheet("📖 Hướng Dẫn")

guide_rows = [
    ("CHRONOFLOW PREMIUM – HƯỚNG DẪN NHẬP DỮ LIỆU TỪ VỰNG", "", ""),
    ("", "", ""),
    ("TÊN TRƯỜNG (FIELD)", "MÔ TẢ CHI TIẾT", "GIÁ TRỊ HỢP LỆ / VÍ DỤ"),
    ("word *",               "Từ tiếng Anh cần học. BẮT BUỘC.",                                           "apple, banana, run, happy"),
    ("meaning *",            "Nghĩa tiếng Việt. BẮT BUỘC. Ngắn gọn, chuẩn tiểu học.",                   "quả táo, chạy, vui vẻ"),
    ("ipa",                  "Phiên âm IPA quốc tế. Bọc trong dấu / /.",                                 "/ˈæp.əl/, /rʌn/"),
    ("vietnamesePhonetic",   "Hướng dẫn đọc bằng tiếng Việt cho bé. Dùng dấu - phân vần.",              "Áp-pờ-lơ, Răn"),
    ("type",                 "Loại từ trong tiếng Anh.",                                                   "Danh từ | Động từ | Tính từ | Trạng từ | Số đếm | Giới từ"),
    ("image",                "Emoji minh họa Unicode hoặc để trống.",                                     "🍎 🐘 🏃 😊 💻"),
    ("hint",                 "Mẹo nhớ từ vui, dưới 80 ký tự. Bắt đầu bằng 💡",                         "💡 Táo đỏ ngọt ngào mỗi ngày!"),
    ("example",              "Câu ví dụ tiếng Anh. Ngắn, phù hợp lứa tuổi 4-10.",                       "I eat an apple."),
    ("exampleVi",            "Bản dịch câu ví dụ sang tiếng Việt.",                                      "Tôi ăn một quả táo."),
    ("level",                "Cấp độ học.",                                                                "L1 (4-5t) | L2 (5-7t) | L3 (7-9t) | L4 (8-10t)"),
    ("category",             "Mã chủ đề bài học. Định dạng: [Level]-U[Số].",                             "L1-U01, L2-U03, L3-U05, L4-U01"),
    ("audioUrl",             "Link file âm thanh MP3. Để trống = dùng Text-to-Speech tự động.",          "https://... hoặc để trống"),
    ("", "", ""),
    ("QUY TẮC NHẬP DỮ LIỆU", "", ""),
    ("1.", "Điền từ hàng 6 trở xuống trong sheet '📥 Nhập Dữ Liệu'.", ""),
    ("2.", "Chỉ 2 cột BẮT BUỘC: word và meaning. Các cột khác tùy chọn.", ""),
    ("3.", "Sau khi điền xong, copy toàn bộ dữ liệu (bỏ hàng tiêu đề).", ""),
    ("4.", "Vào hệ thống → Tab Admin → Import Wizard → Chọn 'Mẫu JSON Chi Tiết' để xem format.", ""),
    ("5.", "Hoặc lưu file dạng CSV rồi dán nội dung vào ô Import Wizard.", ""),
    ("", "", ""),
    ("MÃ CẤP ĐỘ & CHỦ ĐỀ", "", ""),
    ("L1-U01", "Màu sắc (Colors)",            "L3-U01 | Thiên nhiên (Nature)"),
    ("L1-U02", "Số đếm 1-10 (Numbers)",       "L3-U02 | Sức khỏe (Health)"),
    ("L1-U03", "Hình dạng (Shapes)",          "L3-U03 | Mua sắm (Shopping)"),
    ("L1-U04", "Gia đình (My Family)",        "L3-U04 | Du lịch (Travel)"),
    ("L1-U05", "Cơ thể (My Body)",            "L3-U05 | Công nghệ (Technology)"),
    ("L1-U06", "Động vật (Animals)",          "L4-U01 | Không gian (Space)"),
    ("L2-U01", "Đồ ăn & Thức uống (Food)",   "L4-U07 | Quốc gia & Văn hóa"),
    ("L2-U02", "Lớp học (My Classroom)",      "L4-U08 | An toàn số (Digital Safety)"),
    ("L2-U03", "Động từ hành động (Actions)", "L4-U09 | Kể chuyện (Storytelling)"),
    ("L2-U04", "Cảm xúc (Feelings)",         "L4-U10 | Mục tiêu & Trưởng thành"),
]

ws2.column_dimensions["A"].width = 24
ws2.column_dimensions["B"].width = 55
ws2.column_dimensions["C"].width = 42

for r_idx, (a, b, c) in enumerate(guide_rows, start=1):
    ca = ws2.cell(row=r_idx, column=1, value=a)
    cb = ws2.cell(row=r_idx, column=2, value=b)
    cc = ws2.cell(row=r_idx, column=3, value=c)

    if r_idx == 1:
        ws2.merge_cells("A1:C1")
        ca.fill = PatternFill("solid", fgColor="1E293B")
        ca.font = Font(color="FBBF24", bold=True, size=14, name="Calibri")
        ca.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 36
    elif a in ("TÊN TRƯỜNG (FIELD)", "QUY TẮC NHẬP DỮ LIỆU", "MÃ CẤP ĐỘ & CHỦ ĐỀ"):
        for cx in [ca, cb, cc]:
            cx.fill = PatternFill("solid", fgColor="0F766E")
            cx.font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
            cx.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[r_idx].height = 22
    elif a.startswith(("1.", "2.", "3.", "4.", "5.")):
        for cx in [ca, cb, cc]:
            cx.fill = PatternFill("solid", fgColor="EFF6FF")
            cx.font = Font(color="1E40AF", size=10, name="Calibri")
            cx.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws2.row_dimensions[r_idx].height = 20
    elif a.startswith("L") and "-U" in a:
        alt = "F0FDF4" if r_idx % 2 == 0 else "ECFDF5"
        for cx in [ca, cb, cc]:
            cx.fill = PatternFill("solid", fgColor=alt)
            cx.font = Font(color="065F46", size=10, name="Calibri")
            cx.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[r_idx].height = 19
    else:
        alt = "F8FAFC" if r_idx % 2 == 0 else "FFFFFF"
        for cx in [ca, cb, cc]:
            cx.fill = PatternFill("solid", fgColor=alt)
            cx.font = Font(color="1E293B", size=10, name="Calibri")
            cx.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws2.row_dimensions[r_idx].height = 20
        if a and a not in ("", "QUY TẮC NHẬP DỮ LIỆU", "MÃ CẤP ĐỘ & CHỦ ĐỀ"):
            ca.font = Font(color="0F766E", bold=True, size=10, name="Calibri")

# ── Sheet 3: JSON MẪU ─────────────────────────────────────────────────────────
ws3 = wb.create_sheet("📄 JSON Mẫu")
ws3.column_dimensions["A"].width = 100

json_sample = '''[
  {
    "word": "apple",
    "ipa": "/ˈæp.əl/",
    "vietnamesePhonetic": "Áp-pờ-lơ",
    "meaning": "quả táo",
    "type": "Danh từ",
    "image": "🍎",
    "hint": "💡 Táo đỏ ngọt ngào mỗi ngày giúp bé khỏe mạnh!",
    "example": "I eat a red apple every day.",
    "exampleVi": "Tôi ăn một quả táo đỏ mỗi ngày.",
    "level": "L1",
    "category": "L1-U05",
    "audioUrl": ""
  },
  {
    "word": "elephant",
    "ipa": "/ˈel.ə.fənt/",
    "vietnamesePhonetic": "E-lơ-phơn-tơ",
    "meaning": "con voi",
    "type": "Danh từ",
    "image": "🐘",
    "hint": "💡 Con voi có cái vòi rất dài và to lớn nhất!",
    "example": "The elephant drinks water with its trunk.",
    "exampleVi": "Con voi uống nước bằng cái vòi của nó.",
    "level": "L2",
    "category": "L2-U01",
    "audioUrl": ""
  },
  {
    "word": "rainbow",
    "ipa": "/ˈreɪnboʊ/",
    "vietnamesePhonetic": "Rên-bâu",
    "meaning": "cầu vồng",
    "type": "Danh từ",
    "image": "🌈",
    "hint": "💡 7 màu cầu vồng xuất hiện sau cơn mưa rào!",
    "example": "Look at the beautiful rainbow in the sky!",
    "exampleVi": "Hãy nhìn cầu vồng đẹp trên bầu trời!",
    "level": "L3",
    "category": "L3-U02",
    "audioUrl": ""
  }
]

=== HƯỚNG DẪN DÙNG JSON ===
1. Copy toàn bộ nội dung JSON ở trên (từ dấu [ đến dấu ])
2. Vào hệ thống → Tab ADMIN → Import Wizard
3. Dán vào ô "Dán Nội Dung Dữ Liệu Tệp"
4. Bấm "Dry-Run Kiểm Tra" trước để xem trước kết quả
5. Bấm "Thực Thi Nạp Dữ Liệu" để lưu vào CSDL'''

ws3.merge_cells("A1:A1")
t = ws3.cell(row=1, column=1, value="📄 JSON MẪU NHẬP DỮ LIỆU – CHRONOFLOW PREMIUM")
t.fill = PatternFill("solid", fgColor="1E293B")
t.font = Font(color="FBBF24", bold=True, size=14, name="Calibri")
t.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 36

for i, line in enumerate(json_sample.split('\n'), start=2):
    c = ws3.cell(row=i, column=1, value=line)
    is_key = '": "' in line or '": [' in line
    c.font = Font(
        color="22D3EE" if line.strip().startswith('"word"') or line.strip().startswith('"meaning"') else
              "A78BFA" if is_key else
              "34D399" if line.strip().startswith('"') and '":' not in line else
              "F1F5F9",
        size=10, name="Courier New"
    )
    c.fill = PatternFill("solid", fgColor="0F172A")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws3.row_dimensions[i].height = 18

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = r"d:\TÀI LIỆU HỌC CODE\lich-sinh-hoat-react-node-tailwind\CHRONOFLOW_Mau_Nhap_Tu_Vung.xlsx"
wb.save(out_path)
print(f"✅ File Excel đã được tạo tại:\n{out_path}")
